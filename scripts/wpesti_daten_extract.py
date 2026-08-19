# -*- coding: utf-8 -*-
"""
WPesti-Waermepumpen-Datenbank extrahieren (Blatt WP_Daten) -> gema_wpesti_daten.js

Aufruf:  python3 scripts/wpesti_daten_extract.py <WPesti_de.xlsx> [ziel.js]

Quelle: WPesti (EnergieSchweiz/FWS), Blatt WP_Daten — pro Zeile ein Geraet:
  A=Nr (Steuerung)  B=Zeilen_Nr  C=WP-Art (2=Luft-Wasser, 3=Sole-Wasser, 4=Wasser-Wasser)
  D=Hersteller  E=Typ  F=Stufigkeit (1=einstufig / 2=2stufig / 3=mehrstufig / 4=stufenlos)
  G..V  = Luft:   Q/COP je (L-15, L-7, L2, L7, L20)/W35  +  Q/COP je (L-7, L7, L20)/W55
  W..Z  = Sole:   Q B0/W35, COP B0/W35, Q B0/W55, COP B0/W55
  AA..AD= Wasser: Q W10/W35, COP W10/W35, Q W10/W55, COP W10/W55

Zeile 10 ("Eigene Werte") ist der Manuell-Modus des Originals und wird uebersprungen.
Leere Zellen -> 0 (entspricht der OFFSET-Semantik des Originals; die Engine behandelt 0 als fehlend).
"""
import sys, json, datetime

def zahl(v):
    if v is None or v == '':
        return 0
    try:
        f = float(v)
    except (TypeError, ValueError):
        return 0
    if f != f or f in (float('inf'), float('-inf')):
        return 0
    return round(f, 6)

def main():
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)
    src = sys.argv[1]
    ziel = sys.argv[2] if len(sys.argv) > 2 else 'gema_wpesti_daten.js'
    import openpyxl
    wb = openpyxl.load_workbook(src, data_only=True)
    ws = wb['WP_Daten']

    # Version = letzter Eintrag in log-Spalte D
    version = ''
    try:
        log = wb['log']
        for row in log.iter_rows(min_row=5, max_col=4):
            d = row[3].value
            if isinstance(d, str) and d.strip():
                version = d.strip()
    except Exception:
        pass

    lw, sw, ww = {}, {}, {}
    anz = {2: 0, 3: 0, 4: 0}
    uebersprungen = 0
    for row in ws.iter_rows(min_row=11, max_col=30):
        c = [x.value for x in row]
        art = c[2]
        hersteller = (c[3] or '').strip() if isinstance(c[3], str) else c[3]
        typ = (c[4] or '').strip() if isinstance(c[4], str) else c[4]
        if art not in (2, 3, 4) or not hersteller or typ in (None, ''):
            continue
        typ = str(typ)
        stuf = int(c[5]) if isinstance(c[5], (int, float)) and c[5] else 0
        if art == 2:
            werte = [zahl(x) for x in c[6:22]]     # G..V (16)
            ziel_map = lw
        elif art == 3:
            werte = [zahl(x) for x in c[22:26]]    # W..Z (4)
            ziel_map = sw
        else:
            werte = [zahl(x) for x in c[26:30]]    # AA..AD (4)
            ziel_map = ww
        if not any(werte):
            uebersprungen += 1                      # Geraet ohne einen einzigen Messwert
            continue
        liste = ziel_map.setdefault(hersteller, [])
        eintrag = [typ, stuf] + werte
        # Doppelte Typ-Namen: identische Zeile still deduplizieren,
        # abweichende Werte mit « (2)»/« (3)» erreichbar halten (Picker matcht über den Typ-Namen)
        gleiche = [e for e in liste if e[0] == typ or e[0].startswith(typ + ' (')]
        if gleiche:
            if any(e[1:] == eintrag[1:] for e in gleiche):
                continue
            eintrag[0] = typ + ' (' + str(len(gleiche) + 1) + ')'
        liste.append(eintrag)
        anz[art] += 1

    def js_obj(m):
        teile = []
        for h in sorted(m.keys(), key=lambda s: s.lower()):
            zeilen = ',\n  '.join(json.dumps(g, ensure_ascii=False, separators=(',', ':')) for g in m[h])
            teile.append(json.dumps(h, ensure_ascii=False) + ':[\n  ' + zeilen + '\n ]')
        return '{\n ' + ',\n '.join(teile) + '\n}'

    kopf = (
        "/* GEMA — WPesti-Wärmepumpen-Datenbank (Blatt WP_Daten)\n"
        "   AUTOMATISCH ERZEUGT von scripts/wpesti_daten_extract.py — NIE von Hand bearbeiten.\n"
        "   Quelle: WPesti (EnergieSchweiz/FWS), Version " + (version or 'unbekannt') + " · extrahiert " + datetime.date.today().isoformat() + "\n"
        "   lw: Luft-Wasser  [Typ, Stufigkeit, Q-15,COP-15, Q-7,COP-7, Q2,COP2, Q7,COP7, Q20,COP20 (je W35), Q-7,COP-7, Q7,COP7, Q20,COP20 (je W55)]\n"
        "   sw: Sole-Wasser  [Typ, Stufigkeit, Q B0/W35, COP B0/W35, Q B0/W55, COP B0/W55]\n"
        "   ww: Wasser-Wasser[Typ, Stufigkeit, Q W10/W35, COP W10/W35, Q W10/W55, COP W10/W55]\n"
        "   Stufigkeit: 1=einstufig · 2=zweistufig · 3=mehrstufig · 4=stufenlos (reine Anzeige — geht in keine Rechnung ein).\n"
        "   0 = kein Messwert vorhanden. */\n"
    )
    api = """
var _API = (function(){
  var STUFEN = {1:'einstufig', 2:'zweistufig', 3:'mehrstufig', 4:'stufenlos'};
  function gruppe(art){ return (art===3||art===5) ? _SW : (art===4 ? _WW : (art===2 ? _LW : null)); }
  function herstellerListe(art){
    var g = gruppe(art); if(!g) return [];
    var out = []; for(var h in g) if(Object.prototype.hasOwnProperty.call(g,h)) out.push(h);
    out.sort(function(a,b){ return a.toLowerCase() < b.toLowerCase() ? -1 : 1; });
    return out;
  }
  function typen(art, hersteller){
    var g = gruppe(art); if(!g || !g[hersteller]) return [];
    return g[hersteller].map(function(e){ return e[0]; });
  }
  function geraet(art, hersteller, typ){
    var g = gruppe(art); if(!g || !g[hersteller]) return null;
    for(var i=0;i<g[hersteller].length;i++){
      var e = g[hersteller][i];
      if(e[0] === typ){
        var basis = { hersteller:hersteller, typ:e[0], stufigkeit:e[1], stufigkeitName:STUFEN[e[1]]||'' };
        if(g === _LW){
          basis.kennlinie = {
            q35:[e[2],e[4],e[6],e[8],e[10]], cop35:[e[3],e[5],e[7],e[9],e[11]],
            q55:[e[12],e[14],e[16]],         cop55:[e[13],e[15],e[17]]
          };
        } else {
          basis.qB035=e[2]; basis.copB035=e[3]; basis.qB055=e[4]; basis.copB055=e[5];
        }
        return basis;
      }
    }
    return null;
  }
  function anzahl(){ var n=0, g, h; var gs=[_LW,_SW,_WW]; for(var i=0;i<gs.length;i++){ g=gs[i]; for(h in g) if(Object.prototype.hasOwnProperty.call(g,h)) n += g[h].length; } return n; }
  return { version:_V, lw:_LW, sw:_SW, ww:_WW, STUFEN:STUFEN,
           herstellerListe:herstellerListe, typen:typen, geraet:geraet, anzahl:anzahl };
})();
if (typeof window !== 'undefined') window.GemaWpestiDaten = _API;
if (typeof module !== 'undefined' && module.exports) module.exports = _API;
})();
"""
    js = (kopf + '(function(){\n\'use strict\';\n'
          + 'var _V=' + json.dumps(version) + ';\n'
          + 'var _LW=' + js_obj(lw) + ';\n'
          + 'var _SW=' + js_obj(sw) + ';\n'
          + 'var _WW=' + js_obj(ww) + ';\n'
          + api)
    with open(ziel, 'w', encoding='utf-8') as f:
        f.write(js)
    print('Version:', version or 'unbekannt')
    print('Geraete: Luft-Wasser %d · Sole-Wasser %d · Wasser-Wasser %d · total %d (ohne Messwerte uebersprungen: %d)'
          % (anz[2], anz[3], anz[4], sum(anz.values()), uebersprungen))
    print('geschrieben:', ziel)

if __name__ == '__main__':
    main()
